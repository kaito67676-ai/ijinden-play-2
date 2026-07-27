#!/usr/bin/env python3
"""
イジンデン カードデータ ビルドスクリプト
====================================
Excel のカードリスト（第1〜6弾）から、アプリが読み込む cards-data.js を生成する。

元の Excel には「カードの種類」列が無いため、パワー欄・レベル欄・ルールテキストから
イジン / マホウ / ハイケイ / マリョク を推定している。推定しきれないもの、または
元データ自体に誤りがあるものは、下の OVERRIDES 表で明示的に上書きする。
今後カードの誤りを見つけたら、この表に1行足すだけで全体に反映される。
"""

import json
import re
import glob
import os

SRC_DIR = '/mnt/user-data/uploads'
OUT_JSON = '/home/claude/cards.json'
OUT_JS = '/home/claude/ijinden-app/cards-data.js'

FILES = [
    ('1st', 'イシ_ンテ_ン1stlist_230719.xlsx'),
    ('2nd', 'イシ_ンテ_ン2ndlist_231212.xlsx'),
    ('3rd', 'イシ_ンテ_ン3rdlist_240606.xlsx'),
    ('4th', 'イシ_ンテ_ン4thlist_241225.xlsx'),
    ('5th', 'イシ_ンテ_ン5thlist_25091902.xlsx'),
    ('6th', 'イシ_ンテ_ン6thlist_260708.xlsx'),
]

# ============================================================
# 手動修正テーブル（元データの誤り・推定できないものはここで直す）
# ============================================================

# カードの種類の上書き   id -> 種類
TYPE_OVERRIDES = {
    '6th-47': 'ハイケイ',   # 払暁の城壁：持続効果を持つハイケイ。「冥府発動」の語を含むためマホウと誤推定されていた
    '3rd-59': 'ハイケイ',   # 森閑たる離宮：「これが戦場にある間」の持続効果。同じ理由で誤推定されていた

    # 装備テキストを持つ乗り物系はハイケイ。
    # 「魔力ゾーンに置かれているこれを装備させてもよい」と書かれたものはマリョクだが、
    # 下の4枚はその一文が無く、戦場に置かれて持続効果を発揮するハイケイ。
    # （マリョクの目印である遺業能力「復元」も持たず、レベルも5〜7と高い）
    '2nd-50': 'ハイケイ',   # 安宅船
    '2nd-52': 'ハイケイ',   # ガレオン船
    '2nd-54': 'ハイケイ',   # 蒸気機関車
    '3rd-60': 'ハイケイ',   # 黒船
}

# レアリティの上書き   id -> レアリティ
RARITY_OVERRIDES = {
    '3rd-P-16': 'N',   # パープルストーン：元データが空欄
    '6th-2': 'SR',     # 前田慶次：元データが 'm'
}

# 色の上書き   id -> 色
COLOR_OVERRIDES = {}

# ============================================================
# 推定ロジック
# ============================================================

# マリョクを示す手がかり（魔力ゾーンに置かれる／復元を持つ／装備 など）
MARYOKU_SIGNALS = [
    '魔力ゾーンに置かれた', '装備させてもよい', '復元', 'デッキに何枚でも入れてよい',
    '魔力ゾーンにある間', 'これを魔力ゾーンに', '魔力ゾーンで',
    '装備(', '装備：', '装備 :', '装備 ：',
]

# ハイケイを示す手がかり（戦場に置かれ続けて持続的に効果を発揮する言い回し）
HAIKEI_SIGNALS = [
    'これが戦場にある間', 'これが戦場にあるかぎり', 'これが戦場にある限り',
    '戦場か墓地で効果を発揮する', '戦場で効果を発揮する',
]


def classify(card):
    """(種類, レベル, 魔力コスト or None) を返す"""
    power = card['power']
    level_raw = card['level']
    text = (card.get('rule_text') or '') + '\n' + (card.get('legacy_ability') or '')
    level_str = str(level_raw)

    def base_level():
        m = re.match(r'\s*(\d+)', level_str)
        return int(m.group(1)) if m else 0

    # パワーが数値ならイジン
    if isinstance(power, (int, float)):
        return 'イジン', level_raw, None

    # マホウ形式A（第1弾）: "4\n魔力コスト2"
    m = re.search(r'魔力コスト\s*(\d+)', level_str)
    if m:
        return 'マホウ', base_level(), int(m.group(1))

    # マホウ形式B（第2弾以降）: "5□□" → レベル5・魔力コストは□の数
    m = re.match(r'^\s*(\d+)\s*([□]+)\s*$', level_str)
    if m:
        return 'マホウ', int(m.group(1)), len(m.group(2))

    if any(s in text for s in MARYOKU_SIGNALS):
        return 'マリョク', base_level(), None

    # 持続効果の言い回しがあればハイケイ（「冥府発動」の語より優先する）
    if any(s in text for s in HAIKEI_SIGNALS):
        return 'ハイケイ', base_level(), None

    # 魔力コスト表記が抜けているマホウの救済（冥府発動を持つのはマホウのみ）
    if '冥府発動' in text:
        return 'マホウ', base_level(), 0

    return 'ハイケイ', base_level(), None


def load_rows():
    import openpyxl
    all_cards = []
    for setname, fname in FILES:
        path = os.path.join(SRC_DIR, fname)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        header = rows[0]
        for r in rows[1:]:
            if r[0] is None:
                continue
            d = dict(zip(header, r))
            all_cards.append({
                'set': setname,
                'no': d.get('No1-') or d.get('No2-') or d.get('No3-')
                      or d.get('No4-') or d.get('No5-') or d.get('2nd2-'),
                'rarity': d.get('レアリティ') or d.get('レアリティ\n'),
                'color': d.get('色'),
                'name': d.get('名称') or d.get('名称（上部）'),
                'level': d.get('レベル'),
                'power': d.get('パワー'),
                'trait': d.get('特性'),
                'rule_text': d.get('ルールテキスト'),
                'legacy_ability': d.get('遺業能力'),
                'illustrator': d.get('イラストレーター') or d.get('Illustlation'),
            })
    return all_cards


def build():
    cards = load_rows()
    applied = {'type': [], 'rarity': [], 'color': []}

    for c in cards:
        c['id'] = f"{c['set']}-{c['no']}"

        t, lvl, cost = classify(c)
        c['type'] = t
        c['level'] = lvl
        if cost is not None:
            c['magic_cost'] = cost

        # 色の正規化（元データに数値や '-' が混じっている）
        if c['color'] in (None, '-', 28) or not isinstance(c['color'], str):
            c['color'] = '無'

        # 「〜サークル」はマリョク（魔力ゾーンに置くカード）
        if 'サークル' in (c['name'] or '') and c['type'] == 'ハイケイ':
            c['type'] = 'マリョク'

    # 手動修正テーブルの適用
    for c in cards:
        cid = c['id']
        if cid in TYPE_OVERRIDES and c['type'] != TYPE_OVERRIDES[cid]:
            applied['type'].append(f"{cid} {c['name']}: {c['type']} -> {TYPE_OVERRIDES[cid]}")
            c['type'] = TYPE_OVERRIDES[cid]
            c.pop('magic_cost', None)
        if cid in RARITY_OVERRIDES and c['rarity'] != RARITY_OVERRIDES[cid]:
            applied['rarity'].append(f"{cid} {c['name']}: {c['rarity']!r} -> {RARITY_OVERRIDES[cid]}")
            c['rarity'] = RARITY_OVERRIDES[cid]
        if cid in COLOR_OVERRIDES and c['color'] != COLOR_OVERRIDES[cid]:
            applied['color'].append(f"{cid} {c['name']}: {c['color']} -> {COLOR_OVERRIDES[cid]}")
            c['color'] = COLOR_OVERRIDES[cid]

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(cards, f, ensure_ascii=False)
    with open(OUT_JS, 'w', encoding='utf-8') as f:
        f.write('window.IJINDEN_CARDS = ')
        json.dump(cards, f, ensure_ascii=False)
        f.write(';')

    from collections import Counter
    print(f"built {len(cards)} cards")
    print("types:  ", dict(Counter(c['type'] for c in cards)))
    print("rarity: ", dict(Counter(c['rarity'] for c in cards)))
    for kind, items in applied.items():
        if items:
            print(f"\n{kind} overrides applied:")
            for i in items:
                print("  ", i)
    bad = [c['id'] for c in cards if c['type'] == 'マホウ' and 'magic_cost' not in c]
    print("\nマホウ without magic_cost:", bad or 'none')
    return cards


if __name__ == '__main__':
    build()
